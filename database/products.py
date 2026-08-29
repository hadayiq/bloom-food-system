import os
import re
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook


class ProductRepository:

    _products_cache = None
    _products_mtime = None
    _opening_cache = None
    _opening_mtime = None

    def __init__(self):
        self.file = os.path.join(
            os.path.dirname(os.path.dirname(__file__)), "inventory.xlsx"
        )

    def _mtime(self):
        try:
            return os.path.getmtime(self.file)
        except OSError:
            return None

    def get_all_products(self):
        mtime = self._mtime()
        if ProductRepository._products_cache is None or ProductRepository._products_mtime != mtime:
            ProductRepository._products_cache = pd.read_excel(
                self.file, sheet_name="Products"
            )
            ProductRepository._products_mtime = mtime
        return ProductRepository._products_cache.copy()

    def _get_opening_table(self):
        mtime = self._mtime()
        if ProductRepository._opening_cache is None or ProductRepository._opening_mtime != mtime:
            ProductRepository._opening_cache = pd.read_excel(
                self.file, sheet_name="Opening_Balance"
            )
            ProductRepository._opening_mtime = mtime
        return ProductRepository._opening_cache

    @classmethod
    def invalidate_cache(cls):
        cls._products_cache = None
        cls._products_mtime = None
        cls._opening_cache = None
        cls._opening_mtime = None

    def get_product_names(self):
        df = self.get_all_products()
        return df["Product_Name"].dropna().astype(str).str.strip().tolist()

    @staticmethod
    def _normalize_product_name(value):
        if value is None:
            return ""
        return re.sub(r"\s+", " ", str(value).strip()).casefold()

    def get_product_id(self, product_name):
        wanted = self._normalize_product_name(product_name)
        if not wanted:
            return None
        df = self.get_all_products()
        names = df["Product_Name"].map(self._normalize_product_name)
        row = df[names == wanted]
        if row.empty:
            return None
        return row.iloc[0]["product_ID"]

    def get_opening_balance(self, product_id):
        df = self._get_opening_table()
        row = df[df["Product_ID"] == product_id]
        if row.empty:
            return 0
        quantity = row.iloc[0]["Quantity"]
        if pd.isna(quantity):
            return 0
        return float(quantity)

    def add_product(self, product_name, unit, opening_balance):
        workbook = load_workbook(self.file)
        products_sheet = workbook["Products"]
        opening_sheet = workbook["Opening_Balance"]
        last_row = products_sheet.max_row
        if last_row == 1:
            number = 1
        else:
            last_code = products_sheet.cell(row=last_row, column=1).value
            if last_code:
                number = int(last_code.replace("CR", "")) + 1
            else:
                number = 1
        product_id = f"CR{number:03d}"
        products_sheet.append([product_id, product_name, unit])
        opening_sheet.append([
            product_id,
            datetime.now().strftime("%Y-%m-%d"),
            opening_balance,
        ])
        workbook.save(self.file)
        workbook.close()
        self.invalidate_cache()
        from utils.refresh_manager import refresh_manager
        refresh_manager.products_changed.emit()

    def update_product(self, product_id, product_name, unit, opening_balance):
        workbook = load_workbook(self.file)
        products_sheet = workbook["Products"]
        opening_sheet = workbook["Opening_Balance"]
        for row in products_sheet.iter_rows(min_row=2):
            if row[0].value == product_id:
                row[1].value = product_name
                row[2].value = unit
                break
        for row in opening_sheet.iter_rows(min_row=2):
            if row[0].value == product_id:
                row[2].value = opening_balance
                break
        workbook.save(self.file)
        workbook.close()
        self.invalidate_cache()
        from utils.refresh_manager import refresh_manager
        refresh_manager.products_changed.emit()

    def delete_product(self, product_id):
        workbook = load_workbook(self.file)
        products_sheet = workbook["Products"]
        opening_sheet = workbook["Opening_Balance"]
        for row in range(2, products_sheet.max_row + 1):
            if products_sheet.cell(row=row, column=1).value == product_id:
                products_sheet.delete_rows(row)
                break
        for row in range(2, opening_sheet.max_row + 1):
            if opening_sheet.cell(row=row, column=1).value == product_id:
                opening_sheet.delete_rows(row)
                break
        workbook.save(self.file)
        workbook.close()
        self.invalidate_cache()
        from utils.refresh_manager import refresh_manager
        refresh_manager.products_changed.emit()

    def product_exists(self, product_name):
        return self.get_product_id(product_name) is not None
