import os
from datetime import datetime

from openpyxl import load_workbook


class BatchRepository:
    """Repository for product batches stored in inventory.xlsx."""

    SHEET_NAME = "Batches"
    HEADERS = [
        "Batch_ID",
        "Product_ID",
        "Batch_Code",
        "Production_Date",
        "Expiry_Date",
        "Opening_Balance",
    ]

    def __init__(self):
        self.file = os.path.join(
            os.path.dirname(os.path.dirname(__file__)), "inventory.xlsx"
        )
        self._ensure_sheet()

    def _ensure_sheet(self):
        workbook = load_workbook(self.file)
        if self.SHEET_NAME not in workbook.sheetnames:
            sheet = workbook.create_sheet(self.SHEET_NAME)
            sheet.append(self.HEADERS)
            workbook.save(self.file)
        else:
            sheet = workbook[self.SHEET_NAME]
            if sheet.max_row == 1 and all(
                sheet.cell(1, i + 1).value is None for i in range(len(self.HEADERS))
            ):
                for i, header in enumerate(self.HEADERS, start=1):
                    sheet.cell(1, i).value = header
                workbook.save(self.file)
        workbook.close()

    def _next_batch_id(self, sheet):
        max_number = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            value = row[0]
            if isinstance(value, str) and value.startswith("BT"):
                try:
                    max_number = max(max_number, int(value[2:]))
                except ValueError:
                    pass
        return f"BT{max_number + 1:05d}"

    def get_batches(self, product_id=None):
        workbook = load_workbook(self.file, data_only=True)
        sheet = workbook[self.SHEET_NAME]
        result = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            if product_id is not None and row[1] != product_id:
                continue
            result.append(
                {
                    "id": row[0],
                    "product_id": row[1],
                    "code": row[2],
                    "production_date": row[3],
                    "expiry_date": row[4],
                    "opening_balance": float(row[5] or 0),
                }
            )

        workbook.close()
        return result

    def get_batch(self, product_id, batch_code):
        wanted = str(batch_code).strip().casefold()
        for batch in self.get_batches(product_id):
            if str(batch["code"]).strip().casefold() == wanted:
                return batch
        return None

    def batch_exists(self, product_id, batch_code):
        return self.get_batch(product_id, batch_code) is not None

    def add_batch(
        self,
        product_id,
        batch_code,
        production_date,
        expiry_date,
        opening_balance,
    ):
        """Add a new batch with zero opening balance and one incoming movement.

        The supplied quantity represents the quantity received for the new
        batch. It is deliberately stored as an incoming transaction instead
        of Batches.Opening_Balance so it cannot be counted twice at product
        level.
        """
        batch_code = str(batch_code).strip()
        quantity = float(opening_balance)
        if not batch_code:
            raise ValueError("كود الباتش مطلوب.")
        if quantity <= 0:
            raise ValueError("كمية الباتش يجب أن تكون أكبر من صفر.")
        if expiry_date < production_date:
            raise ValueError("تاريخ الصلاحية لا يمكن أن يكون قبل تاريخ الإنتاج.")
        if self.batch_exists(product_id, batch_code):
            raise ValueError("كود الباتش موجود بالفعل لهذا الصنف.")

        workbook = load_workbook(self.file)
        batches_sheet = workbook[self.SHEET_NAME]

        products_sheet = workbook["Products"]
        product_name = None
        for row in products_sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == product_id:
                product_name = str(row[1]).strip()
                break
        if product_name is None:
            workbook.close()
            raise ValueError("الصنف غير موجود.")

        transactions_sheet = workbook["Transactions"]
        if transactions_sheet.max_column < 8:
            transactions_sheet.cell(1, 8).value = "Batch_Code"
        elif transactions_sheet.cell(1, 8).value != "Batch_Code":
            transactions_sheet.cell(1, transactions_sheet.max_column + 1).value = "Batch_Code"

        batch_id = self._next_batch_id(batches_sheet)
        batches_sheet.append(
            [
                batch_id,
                product_id,
                batch_code,
                production_date.strftime("%Y-%m-%d"),
                expiry_date.strftime("%Y-%m-%d"),
                0.0,
            ]
        )

        now = datetime.now()
        transaction_id = f"TR{transactions_sheet.max_row:05d}"
        transactions_sheet.append(
            [
                transaction_id,
                now.strftime("%Y-%m-%d"),
                now.strftime("%H:%M:%S"),
                product_name,
                "مشتريات",
                quantity,
                f"وارد باتش جديد: {batch_code}",
                batch_code,
            ]
        )

        workbook.save(self.file)
        workbook.close()
        return batch_id

    def get_opening_balance(self, product_id, batch_code):
        batch = self.get_batch(product_id, batch_code)
        return float(batch["opening_balance"]) if batch else 0.0

    def get_batch_codes(self, product_id):
        return [batch["code"] for batch in self.get_batches(product_id)]
