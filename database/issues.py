import os
from datetime import datetime

from openpyxl import load_workbook

from database.products import ProductRepository
from database.batches import BatchRepository
from database.transactions import TransactionRepository


class IssueRepository:
    """Issue vouchers: one voucher can contain many product/batch lines."""

    HEADERS = ["Issue_No", "Issue_Date", "Representative", "Status"]
    LINE_HEADERS = ["Issue_No", "Line_No", "Product", "Batch_Code", "Expiry_Date", "Quantity"]
    STOCK_HEADERS = ["Representative", "Product", "Batch_Code", "Quantity", "Updated_At"]
    COUNT_HEADERS = ["Issue_No", "Count_Date", "Representative", "Status"]
    COUNT_LINE_HEADERS = ["Issue_No", "Line_No", "Product", "Batch_Code", "Expiry_Date", "Issued_Quantity", "Counted_Quantity"]
    LIQUIDATION_HEADERS = ["Issue_No", "Liquidation_Date", "Representative", "Total_Issued", "Total_Count", "Total_Sold", "Status"]
    LIQUIDATION_LINE_HEADERS = ["Issue_No", "Line_No", "Product", "Batch_Code", "Expiry_Date", "Issued_Quantity", "Counted_Quantity", "Sold_Quantity"]

    def __init__(self):
        self.file = os.path.join(os.path.dirname(os.path.dirname(__file__)), "inventory.xlsx")
        self.product_repo = ProductRepository()
        self.batch_repo = BatchRepository()
        self.transaction_repo = TransactionRepository()
        self._ensure_sheets()

    def _ensure_sheets(self):
        workbook = load_workbook(self.file)
        changed = False
        for name, headers in [
            ("Issue_Headers", self.HEADERS),
            ("Issue_Lines", self.LINE_HEADERS),
            ("Subwarehouse_Stock", self.STOCK_HEADERS),
            ("Count_Headers", self.COUNT_HEADERS),
            ("Count_Lines", self.COUNT_LINE_HEADERS),
            ("Liquidation_Headers", self.LIQUIDATION_HEADERS),
            ("Liquidation_Lines", self.LIQUIDATION_LINE_HEADERS),
        ]:
            if name not in workbook.sheetnames:
                sheet = workbook.create_sheet(name)
                sheet.append(headers)
                changed = True
        if changed:
            workbook.save(self.file)
        workbook.close()

    def get_next_issue_no(self):
        workbook = load_workbook(self.file, data_only=True)
        sheet = workbook["Issue_Headers"]
        max_number = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            try:
                max_number = max(max_number, int(str(row[0]).strip()))
            except (TypeError, ValueError):
                continue
        workbook.close()
        return f"{max_number + 1:06d}"

    def issue_exists(self, issue_no):
        workbook = load_workbook(self.file, data_only=True)
        sheet = workbook["Issue_Headers"]
        exists = any(str(row[0]).strip() == str(issue_no).strip() for row in sheet.iter_rows(min_row=2, values_only=True) if row[0] is not None)
        workbook.close()
        return exists

    def save_issue(self, issue_no, representative, lines):
        issue_no = str(issue_no).strip()
        representative = str(representative).strip()
        if not issue_no:
            raise ValueError("رقم الإذن مطلوب.")
        if self.issue_exists(issue_no):
            raise ValueError("رقم الإذن موجود بالفعل.")
        if not representative:
            raise ValueError("اسم المندوب / العميل مطلوب.")
        if not lines:
            raise ValueError("يجب إدخال كمية واحدة على الأقل.")

        normalized = []
        for line in lines:
            product = str(line.get("product", "")).strip()
            batch_code = str(line.get("batch_code", "")).strip()
            quantity = float(line.get("quantity", 0) or 0)
            if not product or quantity <= 0:
                continue
            product_id = self.product_repo.get_product_id(product)
            if product_id is None:
                raise ValueError(f"الصنف غير موجود: {product}")
            batch = None
            if batch_code:
                batch = self.batch_repo.get_batch(product_id, batch_code)
                if not batch:
                    raise ValueError(f"الباتش {batch_code} غير موجود للصنف {product}.")
            else:
                if self.batch_repo.get_batches(product_id):
                    raise ValueError(f"يجب اختيار باتش للصنف: {product}")
            if not self.transaction_repo.check_stock(product, quantity, batch_code or None):
                target = f"الباتش {batch_code}" if batch_code else "الصنف"
                raise ValueError(f"الكمية المطلوبة أكبر من رصيد {target}: {product}")
            normalized.append({"product": product, "batch_code": batch_code, "expiry_date": batch["expiry_date"] if batch else "", "quantity": quantity})

        if not normalized:
            raise ValueError("يجب إدخال كمية واحدة على الأقل.")

        workbook = load_workbook(self.file)
        headers = workbook["Issue_Headers"]
        lines_sheet = workbook["Issue_Lines"]
        stock_sheet = workbook["Subwarehouse_Stock"]
        transactions = workbook["Transactions"]
        now = datetime.now()
        headers.append([issue_no, now.strftime("%Y-%m-%d"), representative, "مفتوح"])

        for index, line in enumerate(normalized, start=1):
            lines_sheet.append([issue_no, index, line["product"], line["batch_code"], line["expiry_date"], line["quantity"]])
            transaction_id = f"TR{transactions.max_row:05d}"
            transactions.append([transaction_id, now.strftime("%Y-%m-%d"), now.strftime("%H:%M:%S"), line["product"], "صرف للتسليمات", line["quantity"], f"إذن صرف {issue_no} - {representative}", line["batch_code"]])

            existing_row = None
            for row in range(2, stock_sheet.max_row + 1):
                if (str(stock_sheet.cell(row, 1).value).strip() == representative and str(stock_sheet.cell(row, 2).value).strip() == line["product"] and str(stock_sheet.cell(row, 3).value or "").strip().lower() == line["batch_code"].lower()):
                    existing_row = row
                    break
            if existing_row is None:
                stock_sheet.append([representative, line["product"], line["batch_code"], line["quantity"], now.strftime("%Y-%m-%d %H:%M:%S")])
            else:
                current = float(stock_sheet.cell(existing_row, 4).value or 0)
                stock_sheet.cell(existing_row, 4).value = current + line["quantity"]
                stock_sheet.cell(existing_row, 5).value = now.strftime("%Y-%m-%d %H:%M:%S")

        workbook.save(self.file)
        workbook.close()
        from utils.refresh_manager import refresh_manager
        refresh_manager.data_changed.emit()
        refresh_manager.subwarehouse_changed.emit()
        return issue_no

    def get_open_issues(self):
        workbook = load_workbook(self.file, data_only=True)
        headers = workbook["Issue_Headers"]
        result = []
        for row in headers.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            if str(row[3] or "مفتوح") != "مغلق":
                result.append({"issue_no": str(row[0]), "date": row[1], "representative": row[2], "status": row[3] or "مفتوح"})
        workbook.close()
        return result

    def get_closed_issues_for_rep(self, representative):
        workbook = load_workbook(self.file, data_only=True)
        sheet = workbook["Issue_Headers"]
        result = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            if str(row[2] or "").strip() == str(representative).strip() and str(row[3] or "").strip() == "مغلق":
                result.append({"issue_no": str(row[0]), "date": row[1], "representative": row[2], "status": row[3]})
        workbook.close()
        return result

    def get_subwarehouses(self):
        workbook = load_workbook(self.file, data_only=True)
        sheet = workbook["Subwarehouse_Stock"]
        totals = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            rep = str(row[0]).strip()
            qty = float(row[3] or 0)
            totals[rep] = totals.get(rep, 0.0) + qty
        workbook.close()
        return totals

    def get_subwarehouse_stock(self, representative):
        workbook = load_workbook(self.file, data_only=True)
        sheet = workbook["Subwarehouse_Stock"]
        result = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if str(row[0] or "").strip() != str(representative).strip():
                continue
            result.append({"product": row[1], "batch_code": row[2] or "", "quantity": float(row[3] or 0), "updated_at": row[4]})
        workbook.close()
        return result

    def get_issue(self, issue_no):
        issue_no = str(issue_no).strip()
        workbook = load_workbook(self.file, data_only=True)
        headers = workbook["Issue_Headers"]
        lines = workbook["Issue_Lines"]
        header = None
        for row in headers.iter_rows(min_row=2, values_only=True):
            if str(row[0]).strip() == issue_no:
                header = {"issue_no": str(row[0]), "date": row[1], "representative": row[2], "status": row[3] or "مفتوح"}
                break
        if not header:
            workbook.close()
            return None
        result_lines = []
        for row in lines.iter_rows(min_row=2, values_only=True):
            if str(row[0]).strip() != issue_no:
                continue
            result_lines.append({"line_no": row[1], "product": row[2], "batch_code": row[3] or "", "expiry_date": row[4] or "", "quantity": float(row[5] or 0)})
        workbook.close()
        header["lines"] = result_lines
        return header

    def has_count(self, issue_no):
        workbook = load_workbook(self.file, data_only=True)
        sheet = workbook["Count_Headers"]
        found = any(str(row[0]).strip() == str(issue_no).strip() for row in sheet.iter_rows(min_row=2, values_only=True) if row[0] is not None)
        workbook.close()
        return found

    def save_count(self, issue_no, counted_lines):
        issue = self.get_issue(issue_no)
        if not issue:
            raise ValueError("إذن الصرف غير موجود.")
        if self.has_count(issue_no):
            raise ValueError("تم تسجيل جرد لهذا الإذن بالفعل.")
        if not counted_lines:
            raise ValueError("يجب إدخال الجرد.")

        by_key = {}
        for line in counted_lines:
            key = (str(line.get("product", "")).strip(), str(line.get("batch_code", "")).strip().lower())
            qty = float(line.get("counted_quantity", 0) or 0)
            if qty < 0:
                raise ValueError("الجرد لا يمكن أن يكون رقمًا سالبًا.")
            by_key[key] = qty

        workbook = load_workbook(self.file)
        headers = workbook["Count_Headers"]
        lines = workbook["Count_Lines"]
        now = datetime.now()
        headers.append([issue_no, now.strftime("%Y-%m-%d"), issue["representative"], "مسجل"])

        for index, original in enumerate(issue["lines"], start=1):
            key = (str(original["product"]).strip(), str(original["batch_code"]).strip().lower())
            if key not in by_key:
                workbook.close()
                raise ValueError(f"لم يتم إدخال الجرد للصنف: {original['product']}")
            counted = by_key[key]
            if counted > original["quantity"]:
                workbook.close()
                raise ValueError(f"الجرد للصنف {original['product']} لا يمكن أن يتجاوز إذن الصرف.")
            lines.append([issue_no, index, original["product"], original["batch_code"], original["expiry_date"], original["quantity"], counted])

        workbook.save(self.file)
        workbook.close()
        from utils.refresh_manager import refresh_manager
        refresh_manager.data_changed.emit()
        refresh_manager.subwarehouse_changed.emit()
        return issue_no

    def get_count(self, issue_no):
        workbook = load_workbook(self.file, data_only=True)
        headers = workbook["Count_Headers"]
        lines = workbook["Count_Lines"]
        header = None
        for row in headers.iter_rows(min_row=2, values_only=True):
            if str(row[0]).strip() == str(issue_no).strip():
                header = {"issue_no": str(row[0]), "date": row[1], "representative": row[2], "status": row[3] or "مسجل"}
                break
        if not header:
            workbook.close()
            return None
        result = []
        for row in lines.iter_rows(min_row=2, values_only=True):
            if str(row[0]).strip() == str(issue_no).strip():
                result.append({"line_no": row[1], "product": row[2], "batch_code": row[3] or "", "expiry_date": row[4] or "", "issued_quantity": float(row[5] or 0), "counted_quantity": float(row[6] or 0)})
        workbook.close()
        header["lines"] = result
        return header

    def has_liquidation(self, issue_no):
        workbook = load_workbook(self.file, data_only=True)
        sheet = workbook["Liquidation_Headers"]
        found = any(str(row[0]).strip() == str(issue_no).strip() for row in sheet.iter_rows(min_row=2, values_only=True) if row[0] is not None)
        workbook.close()
        return found

    def liquidate_issue(self, issue_no):
        issue_no = str(issue_no).strip()
        issue = self.get_issue(issue_no)
        count = self.get_count(issue_no)
        if not issue:
            raise ValueError("إذن الصرف غير موجود.")
        if issue["status"] == "مغلق":
            raise ValueError("إذن الصرف مغلق بالفعل.")
        if not count:
            raise ValueError("يجب تسجيل الجرد أولًا.")
        if self.has_liquidation(issue_no):
            raise ValueError("تمت تصفية هذا الإذن بالفعل.")

        count_map = {(str(x["product"]).strip(), str(x["batch_code"]).strip().lower()): float(x["counted_quantity"] or 0) for x in count["lines"]}
        normalized = []
        total_issued = total_counted = total_sold = 0.0
        for original in issue["lines"]:
            key = (str(original["product"]).strip(), str(original["batch_code"]).strip().lower())
            if key not in count_map:
                raise ValueError(f"الجرد ناقص للصنف: {original['product']}")
            issued = float(original["quantity"] or 0)
            counted = float(count_map[key] or 0)
            if counted < 0 or counted > issued:
                raise ValueError(f"كمية الجرد غير صحيحة للصنف: {original['product']}")
            sold = issued - counted
            normalized.append({**original, "counted": counted, "sold": sold})
            total_issued += issued
            total_counted += counted
            total_sold += sold

        workbook = load_workbook(self.file)
        liquidation_headers = workbook["Liquidation_Headers"]
        liquidation_lines = workbook["Liquidation_Lines"]
        stock_sheet = workbook["Subwarehouse_Stock"]
        transactions = workbook["Transactions"]
        issue_headers = workbook["Issue_Headers"]
        now = datetime.now()

        for line in normalized:
            if line["counted"] <= 0:
                continue
            found = None
            for row in range(2, stock_sheet.max_row + 1):
                if (str(stock_sheet.cell(row, 1).value or "").strip() == str(issue["representative"]).strip()
                        and str(stock_sheet.cell(row, 2).value or "").strip() == str(line["product"]).strip()
                        and str(stock_sheet.cell(row, 3).value or "").strip().lower() == str(line["batch_code"] or "").strip().lower()):
                    found = row
                    break
            if found is None or float(stock_sheet.cell(found, 4).value or 0) < line["counted"]:
                workbook.close()
                raise ValueError(f"رصيد المندوب غير كافٍ لإرجاع الجرد: {line['product']}")

        liquidation_headers.append([issue_no, now.strftime("%Y-%m-%d"), issue["representative"], total_issued, total_counted, total_sold, "مغلق"])

        for index, line in enumerate(normalized, start=1):
            liquidation_lines.append([issue_no, index, line["product"], line["batch_code"], line["expiry_date"], line["quantity"], line["counted"], line["sold"]])
            counted = line["counted"]
            if counted <= 0:
                continue

            target_row = None
            for row in range(2, stock_sheet.max_row + 1):
                if (str(stock_sheet.cell(row, 1).value or "").strip() == str(issue["representative"]).strip()
                        and str(stock_sheet.cell(row, 2).value or "").strip() == str(line["product"]).strip()
                        and str(stock_sheet.cell(row, 3).value or "").strip().lower() == str(line["batch_code"] or "").strip().lower()):
                    target_row = row
                    break
            current = float(stock_sheet.cell(target_row, 4).value or 0)
            stock_sheet.cell(target_row, 4).value = current - counted
            stock_sheet.cell(target_row, 5).value = now.strftime("%Y-%m-%d %H:%M:%S")

            transaction_id = f"TR{transactions.max_row:05d}"
            transactions.append([
                transaction_id,
                now.strftime("%Y-%m-%d"),
                now.strftime("%H:%M:%S"),
                line["product"],
                "مردودات تسليمات",
                counted,
                f"مرتجع جرد إذن {issue_no} - {issue['representative']}",
                line["batch_code"],
            ])

        for row in issue_headers.iter_rows(min_row=2):
            if str(row[0].value).strip() == issue_no:
                row[3].value = "مغلق"
                break

        workbook.save(self.file)
        workbook.close()
        from utils.refresh_manager import refresh_manager
        refresh_manager.data_changed.emit()
        refresh_manager.subwarehouse_changed.emit()
        return {"issued": total_issued, "counted": total_counted, "sold": total_sold}
