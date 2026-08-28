import os
import re
from datetime import datetime

from openpyxl import load_workbook

from database.products import ProductRepository
from database.batches import BatchRepository


class TransactionRepository:

    TRANSACTION_IN_TYPES = ["إنتاج", "مشتريات", "مردودات مبيعات", "مردودات تسليمات"]
    TRANSACTION_OUT_TYPES = ["صرف للتجزئة", "صرف للتسليمات"]

    def __init__(self):
        self.file = os.path.join(
            os.path.dirname(os.path.dirname(__file__)), "inventory.xlsx"
        )
        self.batch_repo = BatchRepository()
        self._ensure_batch_column()

    def _ensure_batch_column(self):
        """Check the transaction schema without writing to Excel at startup."""
        workbook = load_workbook(self.file, read_only=True, data_only=True)
        sheet = workbook["Transactions"]
        valid = (
            sheet.max_column >= 8
            and str(sheet.cell(1, 8).value or "").strip() == "Batch_Code"
        )
        workbook.close()
        return valid

    @staticmethod
    def _normalize_product_name(value):
        if value is None:
            return ""
        return re.sub(r"\s+", " ", str(value).strip()).casefold()

    def _canonical_product_name(self, product_repo, product_id):
        products = product_repo.get_all_products()
        for _, row in products.iterrows():
            if row["product_ID"] == product_id:
                return str(row["Product_Name"]).strip()
        return None

    @staticmethod
    def _ensure_batch_column_for_write(sheet):
        if sheet.max_column < 8:
            sheet.cell(1, 8).value = "Batch_Code"
        elif str(sheet.cell(1, 8).value or "").strip() != "Batch_Code":
            sheet.cell(1, sheet.max_column + 1).value = "Batch_Code"

    def save_transaction(self, product, transaction_type, quantity, notes, batch_code=None):
        quantity = float(quantity)
        if quantity <= 0:
            raise ValueError("الكمية يجب أن تكون أكبر من صفر.")
        if transaction_type not in self.TRANSACTION_IN_TYPES + self.TRANSACTION_OUT_TYPES:
            raise ValueError("نوع الحركة غير صالح.")

        product_repo = ProductRepository()
        product_id = product_repo.get_product_id(product)
        if product_id is None:
            raise ValueError("الصنف غير موجود.")
        canonical_product = self._canonical_product_name(product_repo, product_id)

        if batch_code:
            batch = self.batch_repo.get_batch(product_id, batch_code)
            if not batch:
                raise ValueError("الباتش المختار غير موجود لهذا الصنف.")
            batch_code = batch["code"]
            if transaction_type in self.TRANSACTION_OUT_TYPES:
                if not self.check_stock(canonical_product, quantity, batch_code):
                    raise ValueError("الكمية المطلوبة أكبر من رصيد الباتش المتاح.")
                if not self.check_stock(canonical_product, quantity):
                    raise ValueError("الكمية المطلوبة أكبر من رصيد الصنف المتاح.")
        elif transaction_type in self.TRANSACTION_OUT_TYPES and not self.check_stock(canonical_product, quantity):
            raise ValueError("الكمية المطلوبة أكبر من رصيد الصنف المتاح.")

        workbook = load_workbook(self.file)
        sheet = workbook["Transactions"]
        self._ensure_batch_column_for_write(sheet)
        transaction_id = f"TR{sheet.max_row:05d}"
        now = datetime.now()
        sheet.append([
            transaction_id,
            now.strftime("%Y-%m-%d"),
            now.strftime("%H:%M:%S"),
            canonical_product,
            transaction_type,
            quantity,
            notes,
            batch_code or "",
        ])
        workbook.save(self.file)
        workbook.close()

    def get_transactions_by_product(self, product, batch_code=None):
        wanted_product = self._normalize_product_name(product)
        wanted_batch = str(batch_code).strip().casefold() if batch_code is not None else None
        workbook = load_workbook(self.file, data_only=True)
        sheet = workbook["Transactions"]
        transactions = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if self._normalize_product_name(row[3]) != wanted_product:
                continue
            if wanted_batch is not None:
                row_batch = row[7] if len(row) > 7 else None
                if str(row_batch or "").strip().casefold() != wanted_batch:
                    continue
            transactions.append(row)
        workbook.close()
        return transactions

    def _calculate(self, transactions, opening_balance):
        total_in = 0.0
        total_out = 0.0
        for row in transactions:
            transaction_type = row[4]
            quantity = float(row[5] or 0)
            if transaction_type in self.TRANSACTION_IN_TYPES:
                total_in += quantity
            elif transaction_type in self.TRANSACTION_OUT_TYPES:
                total_out += quantity
        balance = float(opening_balance) + total_in - total_out
        return total_in, total_out, balance

    def get_batch_balance(self, product, batch_code):
        product_repo = ProductRepository()
        product_id = product_repo.get_product_id(product)
        if product_id is None:
            return 0.0, 0.0, 0.0
        opening = self.batch_repo.get_opening_balance(product_id, batch_code)
        transactions = self.get_transactions_by_product(product, batch_code)
        return self._calculate(transactions, opening)

    def get_product_balance(self, product):
        """Return (opening, incoming, current_balance) for the product."""
        product_repo = ProductRepository()
        product_id = product_repo.get_product_id(product)
        if product_id is None:
            return 0.0, 0.0, 0.0

        opening_balance = float(product_repo.get_opening_balance(product_id) or 0)
        transactions = self.get_transactions_by_product(product)
        total_in, _total_out, balance = self._calculate(transactions, opening_balance)
        return opening_balance, total_in, balance

    def get_inventory_summary(self):
        product_repo = ProductRepository()
        products = product_repo.get_all_products()
        summary = []
        for _, row in products.iterrows():
            product_name = row["Product_Name"]
            total_opening, total_in, balance = self.get_product_balance(product_name)
            total_out = total_opening + total_in - balance
            summary.append([product_name, total_opening, total_in, total_out, balance])
        return summary

    def check_stock(self, product, requested_quantity, batch_code=None):
        requested_quantity = float(requested_quantity)
        if requested_quantity <= 0:
            return False
        if batch_code:
            _, _, balance = self.get_batch_balance(product, batch_code)
        else:
            _, _, balance = self.get_product_balance(product)
        return balance >= requested_quantity

    def update_transaction(self, transaction_id, product, transaction_type, quantity, notes, batch_code=None):
        quantity = float(quantity)
        if quantity <= 0:
            raise ValueError("الكمية يجب أن تكون أكبر من صفر.")
        if transaction_type not in self.TRANSACTION_IN_TYPES + self.TRANSACTION_OUT_TYPES:
            raise ValueError("نوع الحركة غير صالح.")

        product_repo = ProductRepository()
        product_id = product_repo.get_product_id(product)
        if product_id is None:
            raise ValueError("الصنف غير موجود.")
        canonical_product = self._canonical_product_name(product_repo, product_id)

        existing = self.get_transaction_by_id(transaction_id)
        if existing is None:
            raise ValueError("الحركة المطلوب تعديلها غير موجودة.")

        canonical_batch = ""
        if batch_code:
            batch = self.batch_repo.get_batch(product_id, batch_code)
            if not batch:
                raise ValueError("الباتش المختار غير موجود لهذا الصنف.")
            canonical_batch = batch["code"]

        replacement = {
            "id": transaction_id,
            "product": canonical_product,
            "type": transaction_type,
            "quantity": quantity,
            "batch": canonical_batch,
        }
        self._validate_updated_balances(existing, replacement, product_repo)

        workbook = load_workbook(self.file)
        sheet = workbook["Transactions"]
        self._ensure_batch_column_for_write(sheet)
        found = False
        for row in sheet.iter_rows(min_row=2):
            if row[0].value == transaction_id:
                row[3].value = canonical_product
                row[4].value = transaction_type
                row[5].value = quantity
                row[6].value = notes
                sheet.cell(row=row[0].row, column=8).value = canonical_batch
                found = True
                break
        if not found:
            workbook.close()
            raise ValueError("الحركة المطلوب تعديلها غير موجودة.")
        workbook.save(self.file)
        workbook.close()
        from utils.refresh_manager import refresh_manager
        refresh_manager.data_changed.emit()

    def _validate_updated_balances(self, existing, replacement, product_repo):
        affected_products = {
            self._normalize_product_name(existing.get("product")),
            self._normalize_product_name(replacement.get("product")),
        }
        for normalized_product in affected_products:
            if not normalized_product:
                continue
            product_id = product_repo.get_product_id(normalized_product)
            if product_id is None:
                continue
            product_name = self._canonical_product_name(product_repo, product_id)
            opening = float(product_repo.get_opening_balance(product_id) or 0)
            rows = self.get_transactions_by_product(product_name)
            kept_rows = [row for row in rows if row[0] != existing["id"]]
            _in, _out, balance = self._calculate(kept_rows, opening)
            if self._normalize_product_name(replacement["product"]) == normalized_product:
                if replacement["type"] in self.TRANSACTION_IN_TYPES:
                    balance += replacement["quantity"]
                elif replacement["type"] in self.TRANSACTION_OUT_TYPES:
                    balance -= replacement["quantity"]
            if balance < 0:
                raise ValueError(f"تعديل الحركة سيجعل رصيد الصنف سالبًا: {product_name}.")

        affected_batches = {
            (self._normalize_product_name(existing.get("product")), str(existing.get("batch") or "").strip().casefold()),
            (self._normalize_product_name(replacement.get("product")), str(replacement.get("batch") or "").strip().casefold()),
        }
        for normalized_product, normalized_batch in affected_batches:
            if not normalized_product or not normalized_batch:
                continue
            product_id = product_repo.get_product_id(normalized_product)
            if product_id is None:
                continue
            product_name = self._canonical_product_name(product_repo, product_id)
            batch = self.batch_repo.get_batch(product_id, normalized_batch)
            if not batch:
                raise ValueError("الباتش المختار غير موجود لهذا الصنف.")
            batch_code = batch["code"]
            opening = self.batch_repo.get_opening_balance(product_id, batch_code)
            rows = self.get_transactions_by_product(product_name, batch_code)
            kept_rows = [row for row in rows if row[0] != existing["id"]]
            _in, _out, balance = self._calculate(kept_rows, opening)
            if self._normalize_product_name(replacement["product"]) == normalized_product and str(replacement.get("batch") or "").strip().casefold() == normalized_batch:
                if replacement["type"] in self.TRANSACTION_IN_TYPES:
                    balance += replacement["quantity"]
                elif replacement["type"] in self.TRANSACTION_OUT_TYPES:
                    balance -= replacement["quantity"]
            if balance < 0:
                raise ValueError(f"تعديل الحركة سيجعل رصيد الباتش سالبًا: {batch_code}.")

    def _validate_delete_balance(self, existing, product_repo):
        product_name = existing.get("product")
        normalized_product = self._normalize_product_name(product_name)
        if normalized_product:
            product_id = product_repo.get_product_id(normalized_product)
            if product_id is not None:
                canonical_product = self._canonical_product_name(product_repo, product_id)
                opening = float(product_repo.get_opening_balance(product_id) or 0)
                rows = self.get_transactions_by_product(canonical_product)
                kept_rows = [row for row in rows if row[0] != existing["id"]]
                _in, _out, balance = self._calculate(kept_rows, opening)
                if balance < 0:
                    raise ValueError(f"حذف الحركة سيجعل رصيد الصنف سالبًا: {canonical_product}.")

        batch_code = str(existing.get("batch") or "").strip()
        if not batch_code or not normalized_product:
            return
        product_id = product_repo.get_product_id(normalized_product)
        if product_id is None:
            return
        canonical_product = self._canonical_product_name(product_repo, product_id)
        batch = self.batch_repo.get_batch(product_id, batch_code)
        if batch is None:
            raise ValueError("الباتش المرتبط بالحركة غير موجود لهذا الصنف.")
        canonical_batch = batch["code"]
        opening = self.batch_repo.get_opening_balance(product_id, canonical_batch)
        rows = self.get_transactions_by_product(canonical_product, canonical_batch)
        kept_rows = [row for row in rows if row[0] != existing["id"]]
        _in, _out, balance = self._calculate(kept_rows, opening)
        if balance < 0:
            raise ValueError(f"حذف الحركة سيجعل رصيد الباتش سالبًا: {canonical_batch}.")

    def get_transaction_by_id(self, transaction_id):
        workbook = load_workbook(self.file, data_only=True)
        sheet = workbook["Transactions"]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == transaction_id:
                workbook.close()
                return {"id": row[0], "date": row[1], "time": row[2], "product": row[3], "type": row[4], "quantity": row[5], "notes": row[6], "batch": row[7] if len(row) > 7 else ""}
        workbook.close()
        return None

    def delete_transaction(self, transaction_id):
        product_repo = ProductRepository()
        existing = self.get_transaction_by_id(transaction_id)
        if existing is None:
            raise ValueError("الحركة المطلوب حذفها غير موجودة.")
        self._validate_delete_balance(existing, product_repo)
        workbook = load_workbook(self.file)
        sheet = workbook["Transactions"]
        for row in range(2, sheet.max_row + 1):
            if sheet.cell(row=row, column=1).value == transaction_id:
                sheet.delete_rows(row)
                break
        workbook.save(self.file)
        workbook.close()
        from utils.refresh_manager import refresh_manager
        refresh_manager.data_changed.emit()

    def product_has_transactions(self, product_name):
        wanted_product = self._normalize_product_name(product_name)
        workbook = load_workbook(self.file, data_only=True)
        sheet = workbook["Transactions"]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if self._normalize_product_name(row[3]) == wanted_product:
                workbook.close()
                return True
        workbook.close()
        return False
