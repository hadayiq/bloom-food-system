import os
from openpyxl import load_workbook


class BatchRepository:

    def __init__(self):

        self.file = os.path.join(
            os.path.dirname(os.path.dirname(__file__)), "inventory.xlsx"
        )

    # ==========================
    # إنشاء رقم Batch تلقائي
    # ==========================

    def generate_batch_id(self):

        workbook = load_workbook(self.file)

        sheet = workbook["Batches"]

        return f"B{sheet.max_row:04d}"

    # ==========================
    # إضافة Batch جديد
    # ==========================

    def add_batch(self, product_id, batch_no, production_date, expiry_date, quantity):

        workbook = load_workbook(self.file)

        sheet = workbook["Batches"]

        batch_id = self.generate_batch_id()

        sheet.append(
            [batch_id, product_id, batch_no, production_date, expiry_date, quantity]
        )

        workbook.save(self.file)

    # ==========================
    # كل الباتشات الخاصة بصنف
    # ==========================

    def get_batches_by_product(self, product_id):

        workbook = load_workbook(self.file)

        sheet = workbook["Batches"]

        batches = []

        for row in sheet.iter_rows(min_row=2, values_only=True):

            if row[1] == product_id:

                batches.append(row)

        return batches

    # ==========================
    # البحث عن Batch
    # ==========================

    def get_batch(self, batch_id):

        workbook = load_workbook(self.file)

        sheet = workbook["Batches"]

        for row in sheet.iter_rows(min_row=2, values_only=True):

            if row[0] == batch_id:

                return row

        return None
